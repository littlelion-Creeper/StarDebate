"""
file_parser.py — 多格式文件解析器
==================================
支持的文件类型：
- .md / .txt    : 直接读取
- .pdf          : pdfplumber 提取文本
- .docx         : python-docx 提取文本
- .xlsx         : openpyxl 提取文本
- .csv          : csv 模块读取
- .json         : json 模块读取
- .html         : BeautifulSoup 文本提取

对外 API：
| 方法 | 说明 |
|------|------|
| parse(file_path) → dict | 解析文件，返回统一格式 |
| get_text(file_path) → str | 获取纯文本 |
| is_supported(file_path) → bool | 检查文件类型是否支持 |
"""

import os
import csv
import json as json_module

SUPPORTED_EXTENSIONS = {
    ".md": "markdown",
    ".txt": "text",
    ".pdf": "pdf",
    ".docx": "docx",
    ".xlsx": "xlsx",
    ".csv": "csv",
    ".json": "json",
    ".html": "html",
}


class FileParser:
    """多格式文件解析器 — 统一返回文本接口"""

    @staticmethod
    def get_supported_extensions() -> list:
        """返回支持的文件扩展名列表"""
        return list(SUPPORTED_EXTENSIONS.keys())

    @staticmethod
    def is_supported(file_path: str) -> bool:
        """检查文件类型是否受支持"""
        ext = os.path.splitext(file_path)[1].lower()
        return ext in SUPPORTED_EXTENSIONS

    @staticmethod
    def parse(file_path: str) -> dict:
        """解析文件，返回统一格式。

        Returns:
            {
                "success": bool,
                "text": str,           # 纯文本内容
                "title": str,          # 推断的标题
                "total_chars": int,    # 总字符数
                "file_type": str,      # 文件类型标签
                "error": str|None      # 错误信息
            }
        """
        if not os.path.isfile(file_path):
            return {"success": False, "text": "", "title": "",
                    "total_chars": 0, "file_type": "unknown",
                    "error": f"文件不存在: {file_path}"}

        ext = os.path.splitext(file_path)[1].lower()
        fname = os.path.basename(file_path)
        file_type = SUPPORTED_EXTENSIONS.get(ext, "unknown")

        try:
            if ext in (".md", ".txt", ".log"):
                return FileParser._parse_text(file_path, file_type, fname)
            elif ext == ".pdf":
                return FileParser._parse_pdf(file_path, file_type, fname)
            elif ext == ".docx":
                return FileParser._parse_docx(file_path, file_type, fname)
            elif ext == ".xlsx":
                return FileParser._parse_xlsx(file_path, file_type, fname)
            elif ext == ".csv":
                return FileParser._parse_csv(file_path, file_type, fname)
            elif ext == ".json":
                return FileParser._parse_json(file_path, file_type, fname)
            elif ext == ".html":
                return FileParser._parse_html(file_path, file_type, fname)
            else:
                return {"success": False, "text": "", "title": fname,
                        "total_chars": 0, "file_type": file_type,
                        "error": f"不支持的文件类型: {ext}"}
        except Exception as e:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": str(e)}

    @staticmethod
    def get_text(file_path: str) -> str:
        """获取文件纯文本内容（简化接口）"""
        result = FileParser.parse(file_path)
        return result.get("text", "")

    @staticmethod
    def parse_table(file_path: str, max_rows: int = 50) -> dict:
        """解析表格文件，返回结构化行数据。

        Returns: {"success": bool, "rows": [[str,...],...], "headers": [str,...],
                  "title": str, "error": str|None}
        """
        if not os.path.isfile(file_path):
            return {"success": False, "rows": [], "headers": [], "title": os.path.basename(file_path), "error": "文件不存在"}

        ext = os.path.splitext(file_path)[1].lower()
        fname = os.path.basename(file_path)

        if ext == ".xlsx":
            return FileParser._parse_xlsx_table(file_path, fname, max_rows)
        elif ext == ".csv":
            return FileParser._parse_csv_table(file_path, fname, max_rows)
        else:
            # 非表格文件 → 返回空
            return {"success": False, "rows": [], "headers": [], "title": fname, "error": f"不支持表格解析: {ext}"}

    @staticmethod
    def _parse_xlsx_table(file_path: str, fname: str, max_rows: int) -> dict:
        """解析 XLSX 为结构化行数据"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            all_rows = []
            headers = []
            for sheet_name in wb.sheetnames[:3]:  # 最多3个sheet
                ws = wb[sheet_name]
                for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_rows)):
                    cells = [str(c) if c is not None else "" for c in row]
                    if i == 0 and not headers:
                        headers = cells
                    all_rows.append(cells)
            wb.close()
            return {"success": True, "rows": all_rows, "headers": headers or all_rows[0] if all_rows else [],
                    "title": fname, "error": None}
        except ImportError:
            return {"success": False, "rows": [], "headers": [], "title": fname, "error": "openpyxl 未安装"}
        except Exception as e:
            return {"success": False, "rows": [], "headers": [], "title": fname, "error": str(e)}

    @staticmethod
    def _parse_csv_table(file_path: str, fname: str, max_rows: int) -> dict:
        """解析 CSV 为结构化行数据"""
        try:
            rows = []
            encodings = ["utf-8-sig", "utf-8", "gbk"]
            for enc in encodings:
                try:
                    with open(file_path, "r", encoding=enc) as f:
                        reader = csv.reader(f)
                        for i, row in enumerate(reader):
                            if i >= max_rows:
                                break
                            rows.append(row)
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            headers = rows[0] if rows else []
            return {"success": True, "rows": rows, "headers": headers, "title": fname, "error": None}
        except Exception as e:
            return {"success": False, "rows": [], "headers": [], "title": fname, "error": str(e)}

    # ── 各类型解析器 ────────────────────────────────────────

    @staticmethod
    def _parse_text(file_path: str, file_type: str, fname: str) -> dict:
        """解析纯文本文件 (.md/.txt/.log)"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="gbk", errors="replace") as f:
                text = f.read()

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    @staticmethod
    def _parse_pdf(file_path: str, file_type: str, fname: str) -> dict:
        """解析 PDF 文件"""
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text_parts.append(t)
            text = "\n\n".join(text_parts)
        except ImportError:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": "pdfplumber 未安装。请运行: pip install pdfplumber"}
        except Exception as e:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": str(e)}

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    @staticmethod
    def _parse_docx(file_path: str, file_type: str, fname: str) -> dict:
        """解析 DOCX 文件"""
        try:
            from docx import Document
            doc = Document(file_path)
            text_parts = [p.text for p in doc.paragraphs if p.text.strip()]
            text = "\n\n".join(text_parts)
        except ImportError:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": "python-docx 未安装。请运行: pip install python-docx"}
        except Exception as e:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": str(e)}

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    @staticmethod
    def _parse_xlsx(file_path: str, file_type: str, fname: str) -> dict:
        """解析 XLSX 文件（前 50 行）"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"=== {sheet_name} ===")
                row_count = 0
                for row in ws.iter_rows(values_only=True, max_row=50):
                    row_text = "\t".join([str(c) if c is not None else "" for c in row])
                    if row_text.strip():
                        text_parts.append(row_text)
                        row_count += 1
                text_parts.append(f"(共读取 {row_count} 行)")
            text = "\n".join(text_parts)
            wb.close()
        except ImportError:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": "openpyxl 未安装。请运行: pip install openpyxl"}
        except Exception as e:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": str(e)}

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    @staticmethod
    def _parse_csv(file_path: str, file_type: str, fname: str) -> dict:
        """解析 CSV 文件"""
        try:
            text_parts = []
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row in reader:
                    text_parts.append("\t".join(row))
            text = "\n".join(text_parts)
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="gbk", errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    text_parts.append("\t".join(row))
            text = "\n".join(text_parts)
        except Exception as e:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": str(e)}

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    @staticmethod
    def _parse_json(file_path: str, file_type: str, fname: str) -> dict:
        """解析 JSON 文件 — 提取所有文本字段"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json_module.load(f)
            text_parts = FileParser._extract_text_from_dict(data, max_depth=3)
            text = "\n".join(text_parts)
        except Exception as e:
            return {"success": False, "text": "", "title": fname,
                    "total_chars": 0, "file_type": file_type,
                    "error": str(e)}

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    @staticmethod
    def _parse_html(file_path: str, file_type: str, fname: str) -> dict:
        """解析 HTML 文件 — 提取纯文本"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                html = f.read()
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="gbk", errors="replace") as f:
                html = f.read()

        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, "html.parser")
            for tag in soup(["script", "style", "nav", "footer"]):
                tag.decompose()
            text = soup.get_text(separator="\n", strip=True)
        except ImportError:
            # 简易 HTML 标签移除
            import re
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", "\n", text).strip()

        title = FileParser._extract_title(text, fname)
        return {"success": True, "text": text, "title": title,
                "total_chars": len(text), "file_type": file_type,
                "error": None}

    # ── 工具方法 ────────────────────────────────────────────

    @staticmethod
    def _extract_title(text: str, fallback: str) -> str:
        """从文本中提取标题（第一行非空文本或文件名）"""
        lines = text.strip().split("\n")
        for line in lines:
            stripped = line.strip()
            if stripped and len(stripped) <= 80:
                if stripped.startswith("# "):
                    return stripped[2:].strip()
                return stripped[:72]
        return fallback

    @staticmethod
    def _extract_text_from_dict(obj, max_depth: int = 3, depth: int = 0) -> list:
        """递归提取字典中所有字符串值（排除键名）"""
        results = []
        if depth >= max_depth:
            return results
        if isinstance(obj, dict):
            for key, value in obj.items():
                results.extend(FileParser._extract_text_from_dict(value, max_depth, depth + 1))
        elif isinstance(obj, list):
            for item in obj[:20]:  # 最多 20 个元素
                results.extend(FileParser._extract_text_from_dict(item, max_depth, depth + 1))
        elif isinstance(obj, str) and len(obj.strip()) > 3:
            results.append(obj.strip())
        return results
